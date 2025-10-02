import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math, os

# ------------------ user-editable paths ------------------
VISIT_SCHEDULE_XLSX = "/home/ibab/novohackathon/sched_grid_top/soa_visits_with_groups.xlsx"
FORMS_CSV = "/home/ibab/novohackathon/layout/Final_Complete_eCRF_Matrix.csv"
OUTPUT_XLSX = "final_ptd_full.xlsx"
# --------------------------------------------------------

# read inputs
df_visits = pd.read_excel(VISIT_SCHEDULE_XLSX, sheet_name=0)
df_forms = pd.read_csv(FORMS_CSV)

# normalize col names
df_visits.columns = [c.strip() for c in df_visits.columns]
df_forms.columns = [c.strip() for c in df_forms.columns]

# derive visit info
visit_groups = df_visits["Event Group"].dropna().astype(str).tolist()
visit_names  = df_visits["Visit Name"].dropna().astype(str).tolist()
study_weeks  = df_visits["Study Week"].tolist()

# ------------------ workbook setup ------------------
wb = Workbook()
ws = wb.active
ws.title = "Final PTD"

# styles
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

left_cols = ['Form Label', 'Form Name', 'Source']

# ------------------ HEADER (3 rows) ------------------
# Row 1: Event Group
# Row 2: Visit Name
# Row 3: Study Week
for i, lbl in enumerate(left_cols):
    for row in range(1, 4):
        cell = ws.cell(row=row, column=i+1, value=lbl if row == 2 else None)
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

# visit headers start after left columns
col_start = len(left_cols) + 1
n_visits = len(visit_names)

# Row 2: Visit Name
for j, vname in enumerate(visit_names):
    c = col_start + j
    ws.cell(row=2, column=c, value=vname).font = bold
    ws.cell(row=2, column=c).alignment = center
    ws.cell(row=2, column=c).fill = header_fill
    ws.cell(row=2, column=c).border = border

# Row 3: Study Week
for j, wk in enumerate(study_weeks):
    c = col_start + j
    val = "" if pd.isna(wk) else wk
    ws.cell(row=3, column=c, value=val).font = bold
    ws.cell(row=3, column=c).alignment = center
    ws.cell(row=3, column=c).fill = header_fill
    ws.cell(row=3, column=c).border = border

# Row 1: Event Group (merged)
cur_group = None
start_col = col_start
for j, g in enumerate(visit_groups):
    c = col_start + j
    if cur_group is None:
        cur_group, start_col = g, c
    if g != cur_group:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=c-1)
        ws.cell(row=1, column=start_col, value=cur_group).font = bold
        ws.cell(row=1, column=start_col).alignment = center
        ws.cell(row=1, column=start_col).fill = header_fill
        cur_group, start_col = g, c
# finalize last group
ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_start+n_visits-1)
ws.cell(row=1, column=start_col, value=cur_group).font = bold
ws.cell(row=1, column=start_col).alignment = center
ws.cell(row=1, column=start_col).fill = header_fill

# ------------------ BLOCKS: Visit Dynamics + Event Window ------------------
cur_row = 4

dynamic_rows = [
    "Visit Dynamics (If Y, then Event should appear based on triggering criteria)",
    "Triggering: Event",
    "Triggering: Form",
    "Triggering: Item = Response (if specific response expected, else leave to accept any entered result)"
]

event_window_rows = [
    "Assign Visit Window",
    "Offset Type (Previous Event, Specific Event, or None)",
    "Offset Days (Planned Visit Date, as calculated from Offset Event)",
    "Day Range - Early",
    "Day Range - Late"
]

sections = [("Visit Dynamic Properties", dynamic_rows),
            ("Event Window Configuration", event_window_rows)]

for section_title, attrs in sections:
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(left_cols))
    st_cell = ws.cell(row=cur_row, column=1, value=section_title)
    st_cell.font = bold
    st_cell.alignment = center
    st_cell.fill = grey_fill
    st_cell.border = border
    cur_row += 1

    for attr in attrs:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(left_cols))
        lbl_cell = ws.cell(row=cur_row, column=1, value=attr)
        lbl_cell.font = bold
        lbl_cell.alignment = left_align
        lbl_cell.fill = grey_fill
        lbl_cell.border = border

        for j in range(n_visits):
            c = col_start + j
            mapped_value = ""

            if attr.startswith("Assign Visit Window"):
                mapped_value = "Y"
            elif attr.startswith("Offset Type") and "Offset Type" in df_visits.columns:
                mapped_value = df_visits.iloc[j]["Offset Type"]
            elif attr.startswith("Offset Days") and "Offset Days" in df_visits.columns:
                mapped_value = df_visits.iloc[j]["Offset Days"]
            elif attr in df_visits.columns:
                mapped_value = df_visits.iloc[j][attr]

            if pd.isna(mapped_value):
                mapped_value = ""
            if isinstance(mapped_value, float) and math.isclose(mapped_value, int(mapped_value)):
                mapped_value = int(mapped_value)

            ws.cell(row=cur_row, column=c, value=mapped_value).alignment = center
            ws.cell(row=cur_row, column=c).border = border

        cur_row += 1
    cur_row += 1  # blank row after each section


# ------------------ FORMS TABLE ------------------
forms_start_row = cur_row
df_forms_filtered = df_forms.copy()

row_cursor = forms_start_row
for _, r in df_forms_filtered.iterrows():
    ws.cell(row=row_cursor, column=1, value=r.get('Form Label', '')).alignment = left_align
    ws.cell(row=row_cursor, column=2, value=r.get('Form Name', '')).alignment = left_align
    ws.cell(row=row_cursor, column=3, value=r.get('Source', '')).alignment = left_align

    for j, vname in enumerate(visit_names):
        c = col_start + j
        val = ""
        if vname in r.index:
            val = r[vname]
        if pd.isna(val): val = ""
        if isinstance(val, float) and math.isclose(val, int(val)):
            val = int(val)
        ws.cell(row=row_cursor, column=c, value=val).alignment = center
    row_cursor += 1

# ------------------ formatting ------------------
for col_idx in range(1, ws.max_column+1):
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value is not None:
            lv = str(cell.value)
            if len(lv) > max_len:
                max_len = len(lv)
    ws.column_dimensions[col_letter].width = max(10, max_len + 2)

ws.freeze_panes = ws.cell(row=forms_start_row, column=col_start)

# save
wb.save(OUTPUT_XLSX)
print("✅ Saved:", OUTPUT_XLSX, "size(bytes):", os.path.getsize(OUTPUT_XLSX))
